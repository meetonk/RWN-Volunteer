import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import JobApplication

from django.db.models.deletion import ProtectedError
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Q

from . import models
from .models import JobApplication, ApplicationStatusUpdate, ApplicationStatus
from .forms import JobApplicationForm, ApplicationStatusForm, ApplicationStatusUpdateForm
from django.shortcuts import get_object_or_404

@login_required
def application_list(request):
    apps = JobApplication.objects.filter(user=request.user)
    return render(request, 'application_list.html', {'applications': apps})
default_statuses = ['Applied', '1st Interview', 'Rejected', 'Offer']

@login_required
def application_create(request):
    if request.method == 'POST':
        form = JobApplicationForm(request.POST, user=request.user)
        if form.is_valid():
            app = form.save(commit=False)
            app.user = request.user
            app.save()
            return redirect('application_list')
    else:
        form = JobApplicationForm(user=request.user)
    return render(request, 'application_form.html', {'form': form})




from django.shortcuts import get_object_or_404

@login_required
def add_status_update(request, app_id):
    app = get_object_or_404(JobApplication, id=app_id, user=request.user)

    if request.method == "POST":
        form = ApplicationStatusUpdateForm(request.POST, user=request.user)
        if form.is_valid():
            status_update = form.save(commit=False)
            status_update.job_application = app
            status_update.save()
            return redirect('application_list')
    else:
        form = ApplicationStatusUpdateForm(user=request.user)

    return render(request, 'status_form.html', {'form': form, 'application': app})


@login_required
def application_delete(request, pk):
    app = get_object_or_404(JobApplication, pk=pk, user=request.user)
    app.delete()
    return redirect('application_list')

@login_required
def status_update_delete(request, pk):
    update = get_object_or_404(ApplicationStatusUpdate, pk=pk)
    job_id = update.job_application.id
    update.delete()
    return redirect('application_list')

@login_required
def status_delete(request, pk):
    status = get_object_or_404(ApplicationStatus, pk=pk, user=request.user)
    try:
        status.delete()
        messages.success(request, "Status deleted successfully.")
    except ProtectedError:
        messages.error(request, "Cannot delete this status because it's used in a status update.")
    return redirect('status_create')


@login_required
def status_create(request):
    if request.method == 'POST':
        form = ApplicationStatusForm(request.POST)
        if form.is_valid():
            status = form.save(commit=False)
            status.user = request.user
            status.save()
            return redirect('status_create')
    else:
        form = ApplicationStatusForm()

    statuses = ApplicationStatus.objects.filter(Q(user=request.user) | Q(name__in=default_statuses))

    return render(request, 'status_form.html', {'form': form, 'statuses': statuses})


def create_default_statuses():
    for status in ["Applied", "1st Interview", "Rejected", "Offer"]:
        ApplicationStatus.objects.get_or_create(name=status, user=None)




def export_applications_excel(request):
    user = request.user
    if not user.is_authenticated:
        return HttpResponse("Unauthorized", status=401)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Write headers
    headers = ["Name", "Position", "Activity"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header

    # Get user's applications
    applications = JobApplication.objects.filter(user=user)

    for row_num, app in enumerate(applications, 2):
        ws.cell(row=row_num, column=1).value = app.company
        ws.cell(row=row_num, column=2).value = app.position

        # Gather status updates ordered by date
        status_updates = app.status_updates.order_by('date')

        # Format status text lines: "Status Name - Date (optional notes)"
        status_lines = []
        for update in status_updates:
            line = f"{update.status.name} - {update.date.strftime('%m/%d/%Y')}"
            if update.notes:
                line += f" ({update.notes})"
            status_lines.append(line)

        # Join all lines with newlines
        statuses_text = "\n".join(status_lines) if status_lines else "No status updates"

        cell = ws.cell(row=row_num, column=3)
        cell.value = statuses_text

        # Enable text wrap in this cell
        cell.alignment = Alignment(wrap_text=True)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        if col == 3:
            # Make status column wider
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=job_applications.xlsx'
    wb.save(response)
    return response